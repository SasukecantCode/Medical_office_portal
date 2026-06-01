from __future__ import annotations

import csv
import io
import os
import json
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Request, Form
from sqlalchemy.orm import Session

from app.core.config import settings
from app.db.session import get_db
from app.crud import hr_field_defs as crud_defs
from app.crud import hr_staff as crud_staff
from app.schemas.hr_field_def import HRFieldDefCreate, HRFieldDefRead, HRFieldDefUpdate
from app.schemas.hr_staff import HRStaffCreate, HRStaffUpdate

router = APIRouter(prefix="/agent")


@router.get("/fields", response_model=list[HRFieldDefRead])
def list_fields(db: Session = Depends(get_db)):
    return crud_defs.list_field_defs(db)


@router.post("/fields", response_model=HRFieldDefRead)
def create_field(payload: HRFieldDefCreate, db: Session = Depends(get_db)):
    return crud_defs.create_field_def(db, payload)


@router.patch("/fields/{field_id}", response_model=HRFieldDefRead)
def patch_field(field_id: int, payload: HRFieldDefUpdate, db: Session = Depends(get_db)):
    f = crud_defs.update_field_def(db, field_id, payload)
    if f is None:
        raise HTTPException(status_code=404, detail="Field definition not found")
    return f


@router.delete("/fields/{field_id}")
def delete_field(field_id: int, db: Session = Depends(get_db)):
    ok = crud_defs.delete_field_def(db, field_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Field definition not found")
    return {"status": "deleted"}


def _read_csv_headers_and_samples(file_bytes: bytes, max_rows: int = 5) -> dict:
    s = io.StringIO(file_bytes.decode("utf-8", errors="ignore"))
    reader = csv.DictReader(s)
    headers = list(reader.fieldnames or [])
    samples = []
    for i, row in enumerate(reader):
        if i >= max_rows:
            break
        samples.append(row)
    return {"headers": headers, "samples": samples}


@router.post("/parse")
def parse_file(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Parse uploaded CSV/XLSX and return headers, samples and simple suggestions.
    Suggestions map incoming headers to existing core fields when names match (case-insensitive).
    """
    data = file.file.read()
    filename = file.filename or "upload"
    lower = filename.lower()
    if lower.endswith(".csv"):
        parsed = _read_csv_headers_and_samples(data)
    else:
        try:
            # openpyxl is optional; provide helpful error if missing
            from openpyxl import load_workbook

            wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                raise HTTPException(status_code=400, detail="Workbook has no active worksheet")
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                headers = []
                samples = []
            else:
                headers = [str(c) if c is not None else "" for c in rows[0]]
                samples = []
                for r in rows[1:6]:
                    samples.append({headers[i] if i < len(headers) else f"col_{i}": (v if v is not None else "") for i, v in enumerate(r)})
            parsed = {"headers": headers, "samples": samples}
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Failed to parse file: {exc}")

    # simple suggestions: map header -> existing field name if direct match
    existing = [f.name for f in crud_defs.list_field_defs(db)]
    core = [
        "id",
        "full_name",
        "gender",
        "date_of_birth",
        "designation",
        "cadre",
        "employment_type",
        "phone",
        "email",
        "facility_name",
        "facility_type",
        "district",
        "block",
        "posting_place",
        "date_of_joining",
        "remarks",
    ]
    known = set(existing + core)

    suggestions: dict[str, Any] = {}
    for h in parsed.get("headers", []):
        if not h:
            continue
        key = h.strip()
        low = key.lower().replace(" ", "_")
        if low in known:
            suggestions[key] = low
        else:
            suggestions[key] = None

    return {"filename": filename, "headers": parsed.get("headers", []), "samples": parsed.get("samples", []), "suggestions": suggestions}


@router.post("/import")
def apply_import(
    file: UploadFile = File(...),
    mapping: str | None = Form(None),
    db: Session = Depends(get_db),
):
    """Apply an import: client provides file and a mapping from file headers -> target field names.
    Unmapped fields are stored under `extra` as-is. The `mapping` is expected as a JSON string in form data.
    """
    # Parse mapping from form field (JSON string)
    try:
        mapping_dict = json.loads(mapping) if mapping else {}
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid mapping JSON")

    data = file.file.read()
    filename = (file.filename or "upload").lower()
    rows: list[dict] = []
    if filename.endswith(".csv"):
        s = io.StringIO(data.decode("utf-8", errors="ignore"))
        reader = csv.DictReader(s)
        for r in reader:
            rows.append(r)
    else:
        try:
            from openpyxl import load_workbook

            wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                raise HTTPException(status_code=400, detail="Workbook has no active worksheet")
            it = ws.iter_rows(values_only=True)
            headers = [str(c) if c is not None else "" for c in next(it)]
            for r in it:
                obj = {headers[i]: (r[i] if i < len(r) else "") for i in range(len(headers))}
                rows.append(obj)
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Failed to parse file: {exc}")

    created = 0
    skipped = 0
    out_items = []
    for r in rows:
        payload = {}
        extra = {}
        for k, v in r.items():
            target = mapping_dict.get(k) if mapping_dict else None
            if target:
                payload[target] = v
            else:
                extra[k] = v

        # Attach extra
        payload["extra"] = extra
        try:
            # Validate minimal required fields
            if not payload.get("full_name") or not payload.get("designation") or not payload.get("facility_name"):
                skipped += 1
                continue
            obj = HRStaffCreate.model_validate(payload)
            staff = crud_staff.create_staff(db, obj)
            created += 1
            out_items.append({"id": staff.id, "full_name": staff.full_name})
        except Exception:
            skipped += 1

    return {"created": created, "skipped": skipped, "items": out_items}


@router.post("/chat")
def chat_endpoint(prompt: dict, request: Request, db: Session = Depends(get_db)):
    """Chat endpoint with two modes:
    - ask: plain Gemini response
    - agent: Gemini + backend tool calls
    """
    # Prefer using configured settings (reads from .env via pydantic)
    key = getattr(settings, "gemini_api_key", None) or os.environ.get("GEMINI_API_KEY") or os.environ.get("LLM_API_KEY")
    if not key:
        return {"reply": "Agent functionality not enabled. Set GEMINI_API_KEY in backend/.env or environment to enable."}

    import httpx

    def _to_contents(messages: list[dict]) -> list[dict]:
        contents: list[dict] = []
        for msg in messages:
            role = msg.get("role", "user")
            content = str(msg.get("content", ""))
            if not content:
                continue
            # Gemini supports user/model roles in conversation contents.
            gemini_role = "model" if role in {"assistant", "model"} else "user"
            contents.append({"role": gemini_role, "parts": [{"text": content}]})
        return contents

    def _extract_text(response: dict) -> str:
        candidates = response.get("candidates", [])
        if not candidates:
            return "No response from assistant."
        parts = candidates[0].get("content", {}).get("parts", [])
        texts = [p.get("text", "") for p in parts if isinstance(p, dict) and p.get("text")]
        joined = "\n".join(t for t in texts if t).strip()
        return joined or "No response from assistant."

    def _call_gemini(body: dict) -> dict:
        model = getattr(settings, "gemini_model", None) or os.environ.get("GEMINI_MODEL") or "gemini-2.5-pro"
        model = model.removeprefix("models/")
        url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={key}"
        headers = {"Content-Type": "application/json"}
        r = httpx.post(url, headers=headers, json=body, timeout=30.0)
        r.raise_for_status()
        return r.json()

    def _tool_list_staff(args: dict) -> dict:
        limit = int(args.get("limit", 10) or 10)
        limit = max(1, min(limit, 50))
        rows = crud_staff.list_staff(
            db,
            skip=0,
            limit=limit,
            q=args.get("q"),
            district=args.get("district"),
            designation=args.get("designation"),
            facility_name=args.get("facility_name"),
            employment_type=args.get("employment_type"),
        )
        return {
            "count": len(rows),
            "items": [
                {
                    "id": r.id,
                    "full_name": r.full_name,
                    "designation": r.designation,
                    "facility_name": r.facility_name,
                    "district": r.district,
                    "phone": r.phone,
                    "email": r.email,
                }
                for r in rows
            ],
        }

    def _tool_bulk_id_cards(args: dict) -> dict:
        limit = int(args.get("limit", 500) or 500)
        limit = max(1, min(limit, 500))
        rows = crud_staff.list_staff(
            db,
            skip=0,
            limit=limit,
            q=args.get("q"),
            district=args.get("district"),
            designation=args.get("designation"),
            facility_name=args.get("facility_name"),
            employment_type=args.get("employment_type"),
        )
        base_url = str(request.base_url).rstrip("/") if request else ""

        def _staff_to_card_dict(s) -> dict:
            photo_url = ""
            if getattr(s, "profile_photo_stored_filename", None):
                photo_url = f"{base_url}/api/hr/staff/{s.id}/photo"
            return {
                "id": s.id,
                "full_name": s.full_name,
                "gender": s.gender,
                "designation": s.designation,
                "cadre": s.cadre,
                "employment_type": s.employment_type,
                "facility_name": s.facility_name,
                "facility_type": s.facility_type,
                "posting_place": s.posting_place,
                "block": s.block,
                "district": s.district,
                "date_of_birth": s.date_of_birth.isoformat() if s.date_of_birth else None,
                "date_of_joining": s.date_of_joining.isoformat() if s.date_of_joining else None,
                "phone": s.phone,
                "email": s.email,
                "remarks": s.remarks,
                "extra": s.extra or {},
                "photo_url": photo_url,
            }

        field_defs = crud_defs.list_field_defs(db)
        return {
            "count": len(rows),
            "items": [_staff_to_card_dict(r) for r in rows],
            "field_defs": [
                {
                    "id": f.id,
                    "name": f.name,
                    "label": f.label,
                    "data_type": f.data_type,
                    "sort_order": f.sort_order,
                    "required": f.required,
                }
                for f in field_defs
            ],
            "filters": {
                "q": args.get("q"),
                "district": args.get("district"),
                "designation": args.get("designation"),
                "facility_name": args.get("facility_name"),
                "employment_type": args.get("employment_type"),
                "limit": limit,
            },
        }

    def _tool_export_staff(args: dict) -> dict:
        fmt = str(args.get("format", "xlsx") or "xlsx").lower()
        if fmt not in {"csv", "xlsx"}:
            fmt = "xlsx"
        base_url = str(request.base_url).rstrip("/") if request else ""
        qs_parts = [f"format={fmt}"]
        for key in ["q", "district", "designation", "facility_name", "employment_type"]:
            val = args.get(key)
            if val:
                qs_parts.append(f"{key}={val}")
        qs_str = "&".join(qs_parts)
        download_url = f"{base_url}/api/hr/staff/export?{qs_str}"
        return {
            "status": "ready",
            "format": fmt,
            "download_url": download_url,
        }

    def _tool_navigate_page(args: dict) -> dict:
        valid_pages = {
            "dashboard-home", "staff-list", "add-staff",
            "edit-staff", "id-cards", "attachments",
        }
        page = str(args.get("page", "") or "").strip()
        if page not in valid_pages:
            return {"error": f"Unknown page: {page}. Valid: {', '.join(sorted(valid_pages))}"}
        return {"status": "navigated", "page": page}

    def _tool_get_staff(args: dict) -> dict:
        staff_id = int(args.get("staff_id", 0) or 0)
        if staff_id <= 0:
            return {"error": "staff_id is required"}
        s = crud_staff.get_staff(db, staff_id)
        if not s:
            return {"error": f"Staff with id {staff_id} not found"}
        return {
            "id": s.id,
            "full_name": s.full_name,
            "designation": s.designation,
            "facility_name": s.facility_name,
            "district": s.district,
            "phone": s.phone,
            "email": s.email,
            "cadre": s.cadre,
            "employment_type": s.employment_type,
            "remarks": s.remarks,
            "extra": s.extra or {},
        }

    def _tool_dashboard_summary(_: dict) -> dict:
        return crud_staff.dashboard_summary(db)

    def _tool_list_field_defs(_: dict) -> dict:
        defs = crud_defs.list_field_defs(db)
        return {
            "count": len(defs),
            "items": [
                {
                    "id": f.id,
                    "name": f.name,
                    "label": f.label,
                    "data_type": f.data_type,
                    "sort_order": f.sort_order,
                    "required": f.required,
                }
                for f in defs
            ],
        }

    def _compute_sort_order(args: dict) -> int:
        """Compute sort_order based on insert_after or position hint."""
        insert_after = args.pop("insert_after", None)
        position = args.pop("position", None)

        defs = crud_defs.list_field_defs(db)

        if insert_after:
            needle = str(insert_after).strip().lower().replace(" ", "_")
            for f in defs:
                if f.name.lower() == needle or f.label.lower().replace(" ", "_") == needle:
                    later = [d.sort_order for d in defs if d.sort_order > f.sort_order]
                    if later:
                        return (f.sort_order + min(later)) // 2 or f.sort_order + 1
                    return f.sort_order + 10
            return (max((d.sort_order for d in defs), default=0)) + 10

        if position == "first":
            return (min((d.sort_order for d in defs), default=0)) - 10

        if "sort_order" in args:
            return int(args["sort_order"])

        return (max((d.sort_order for d in defs), default=0)) + 10

    def _tool_create_field_def(args: dict) -> dict:
        try:
            args["sort_order"] = _compute_sort_order(args)
            payload = HRFieldDefCreate.model_validate(args)
        except Exception as exc:
            return {"error": "invalid_payload", "message": str(exc)}
        created = crud_defs.create_field_def(db, payload)
        return {
            "status": "created",
            "id": created.id,
            "name": created.name,
            "label": created.label,
            "data_type": created.data_type,
            "sort_order": created.sort_order,
            "required": created.required,
        }

    def _tool_update_field_def(args: dict) -> dict:
        field_id = int(args.pop("field_id", 0) or 0)
        if field_id <= 0:
            return {"error": "field_id is required"}
        if "insert_after" in args or "position" in args:
            args["sort_order"] = _compute_sort_order(args)
        try:
            payload = HRFieldDefUpdate.model_validate(args)
        except Exception as exc:
            return {"error": "invalid_payload", "message": str(exc)}
        f = crud_defs.update_field_def(db, field_id, payload)
        if f is None:
            return {"error": f"Field with id {field_id} not found"}
        return {"status": "updated", "id": f.id, "name": f.name, "label": f.label}

    def _tool_delete_field_def(args: dict) -> dict:
        field_id = int(args.get("field_id", 0) or 0)
        if field_id <= 0:
            return {"error": "field_id is required"}
        ok = crud_defs.delete_field_def(db, field_id)
        if not ok:
            return {"error": f"Field with id {field_id} not found"}
        return {"status": "deleted", "id": field_id}

    def _tool_create_staff(args: dict) -> dict:
        try:
            payload = HRStaffCreate.model_validate(args)
        except Exception as exc:
            return {"error": "invalid_payload", "message": str(exc)}
        created = crud_staff.create_staff(db, payload)
        return {"status": "created", "id": created.id, "full_name": created.full_name}

    def _tool_update_staff(args: dict) -> dict:
        staff_id = int(args.pop("staff_id", 0) or 0)
        if staff_id <= 0:
            return {"error": "staff_id is required"}
        try:
            payload = HRStaffUpdate.model_validate(args)
        except Exception as exc:
            return {"error": "invalid_payload", "message": str(exc)}
        s = crud_staff.update_staff(db, staff_id, payload)
        if s is None:
            return {"error": f"Staff with id {staff_id} not found"}
        return {"status": "updated", "id": s.id, "full_name": s.full_name}

    def _tool_delete_staff(args: dict) -> dict:
        staff_id = int(args.get("staff_id", 0) or 0)
        if staff_id <= 0:
            return {"error": "staff_id is required"}
        ok = crud_staff.delete_staff(db, staff_id)
        if not ok:
            return {"error": f"Staff with id {staff_id} not found"}
        return {"status": "deleted", "id": staff_id}

    tools = {
        "list_staff": _tool_list_staff,
        "get_staff": _tool_get_staff,
        "create_staff": _tool_create_staff,
        "update_staff": _tool_update_staff,
        "delete_staff": _tool_delete_staff,
        "dashboard_summary": _tool_dashboard_summary,
        "list_field_defs": _tool_list_field_defs,
        "bulk_id_cards": _tool_bulk_id_cards,
        "create_field_def": _tool_create_field_def,
        "update_field_def": _tool_update_field_def,
        "delete_field_def": _tool_delete_field_def,
        "export_staff": _tool_export_staff,
        "navigate_page": _tool_navigate_page,
    }

    try:
        mode = str(prompt.get("mode", "ask") or "ask").lower()
        if mode not in {"ask", "agent"}:
            mode = "ask"

        messages = prompt.get("messages", [])
        if not messages:
            raise HTTPException(status_code=400, detail="No messages provided")

        contents = _to_contents(messages)
        if not contents:
            raise HTTPException(status_code=400, detail="No usable message content provided")

        if mode == "ask":
            response = _call_gemini({"contents": contents})
            return {"reply": _extract_text(response), "raw": response, "mode": mode}

        # ── Agent mode: full tool access, no write gating ──
        system_instruction = (
            "You are an HR data assistant for a Medical Office Portal. "
            "You have FULL read and write access. Do EVERYTHING a user/admin can do. "
            "Do NOT ask for confirmations.\n\n"
            "=== CAPABILITIES ===\n"
            "1. STAFF: list_staff, get_staff, create_staff, update_staff, delete_staff\n"
            "2. ID CARDS: bulk_id_cards (UI automatically renders & zips). Use when asked to generate cards.\n"
            "3. EXPORT: export_staff(format='xlsx' or 'csv')\n"
            "4. FIELDS: list_field_defs, create_field_def, update_field_def, delete_field_def\n"
            "5. NAVIGATION: navigate_page to dashboard-home, staff-list, add-staff, id-cards\n"
            "6. DASHBOARD: dashboard_summary\n\n"
            "=== RULES ===\n"
            "- Act immediately.\n"
            "- If user says 'make ID cards for everyone', call bulk_id_cards.\n"
            "- If user says 'export data', call export_staff.\n"
        )

        tool_decls = [
            {
                "name": "list_staff",
                "description": "Search staff records.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "q": {"type": "string"},
                        "district": {"type": "string"},
                        "designation": {"type": "string"},
                        "facility_name": {"type": "string"},
                        "employment_type": {"type": "string"},
                        "limit": {"type": "number"},
                    },
                },
            },
            {
                "name": "get_staff",
                "description": "Get full details for one staff record by id.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "staff_id": {"type": "number"},
                    },
                    "required": ["staff_id"],
                },
            },
            {
                "name": "create_staff",
                "description": "Create a new staff record.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "full_name": {"type": "string"},
                        "designation": {"type": "string"},
                        "gender": {"type": "string"},
                        "date_of_birth": {"type": "string"},
                        "facility_name": {"type": "string"},
                        "district": {"type": "string"},
                    },
                    "required": ["full_name"],
                },
            },
            {
                "name": "update_staff",
                "description": "Update an existing staff record.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "staff_id": {"type": "number"},
                        "full_name": {"type": "string"},
                        "designation": {"type": "string"},
                        "facility_name": {"type": "string"},
                        "district": {"type": "string"},
                    },
                    "required": ["staff_id"],
                },
            },
            {
                "name": "delete_staff",
                "description": "Delete a staff record.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "staff_id": {"type": "number"},
                    },
                    "required": ["staff_id"],
                },
            },
            {
                "name": "dashboard_summary",
                "description": "Get dashboard totals and grouped stats for HR staff.",
                "parameters": {
                    "type": "object",
                    "properties": {},
                },
            },
            {
                "name": "list_field_defs",
                "description": "List all custom HR field definitions with their id, name, label, data_type, sort_order, and required status.",
                "parameters": {
                    "type": "object",
                    "properties": {},
                },
            },
            {
                "name": "bulk_id_cards",
                "description": "Generate ID cards in bulk. Returns data for the UI to render and zip.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "q": {"type": "string"},
                        "district": {"type": "string"},
                        "designation": {"type": "string"},
                        "facility_name": {"type": "string"},
                        "employment_type": {"type": "string"},
                        "limit": {"type": "number"},
                    },
                },
            },
            {
                "name": "export_staff",
                "description": "Export staff records as an Excel (.xlsx) or CSV file.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "format": {"type": "string"},
                        "q": {"type": "string"},
                        "district": {"type": "string"},
                        "designation": {"type": "string"},
                        "facility_name": {"type": "string"},
                        "employment_type": {"type": "string"},
                    },
                },
            },
            {
                "name": "navigate_page",
                "description": "Navigate to a specific page in the portal UI: dashboard-home, staff-list, add-staff, id-cards.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "page": {
                            "type": "string",
                        },
                    },
                    "required": ["page"],
                },
            },
            {
                "name": "create_field_def",
                "description": "Create a new custom HR field definition.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "name": {"type": "string"},
                        "label": {"type": "string"},
                        "data_type": {"type": "string"},
                        "required": {"type": "boolean"},
                        "insert_after": {"type": "string"},
                        "position": {"type": "string"},
                    },
                    "required": ["name", "label", "data_type"],
                },
            },
            {
                "name": "update_field_def",
                "description": "Update an existing field definition by id. Can change name, label, data_type, required, or reposition with insert_after/position.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "field_id": {"type": "number", "description": "ID of the field to update"},
                        "name": {"type": "string"},
                        "label": {"type": "string"},
                        "data_type": {"type": "string"},
                        "required": {"type": "boolean"},
                        "insert_after": {"type": "string"},
                        "position": {"type": "string"},
                    },
                    "required": ["field_id"],
                },
            },
            {
                "name": "delete_field_def",
                "description": "Delete a field definition by id.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "field_id": {"type": "number", "description": "ID of the field to delete"},
                    },
                    "required": ["field_id"],
                },
            },
        ]

        executed_tools: list[dict] = []
        for _ in range(6):
            body: dict[str, Any] = {
                "contents": contents,
                "tools": [{"functionDeclarations": tool_decls}],
                "systemInstruction": {"parts": [{"text": system_instruction}]},
            }
            response = _call_gemini(body)

            candidate = (response.get("candidates") or [{}])[0]
            model_content = candidate.get("content") or {}
            parts = model_content.get("parts") or []
            function_calls = [p.get("functionCall") for p in parts if isinstance(p, dict) and p.get("functionCall")]

            if not function_calls:
                return {
                    "reply": _extract_text(response),
                    "raw": response,
                    "mode": mode,
                    "actions": executed_tools,
                }

            # Keep model turn in conversation before sending function responses.
            contents.append(model_content)
            for fc in function_calls:
                if not isinstance(fc, dict):
                    continue
                name = fc.get("name")
                args = fc.get("args") or {}

                if name in tools:
                    result = tools[name](args)
                else:
                    result = {"error": f"Unknown tool: {name}"}

                executed_tools.append({"tool": name, "args": args, "result": result})
                contents.append(
                    {
                        "role": "user",
                        "parts": [
                            {
                                "functionResponse": {
                                    "name": name,
                                    "response": {"result": result},
                                }
                            }
                        ],
                    }
                )

        # Fallback if agent loop exceeds max iterations.
        return {
            "reply": "I reached the tool execution limit. Please refine your request.",
            "mode": mode,
            "actions": executed_tools,
        }

    except httpx.HTTPStatusError as exc:
        status = exc.response.status_code if exc.response is not None else None
        error_body = exc.response.text if exc.response else "Unknown error"
        # Helpful hint for common 404: model name not available in this project/API version
        if status == 404:
            hint = (
                "Model not found (404). Update backend/.env GEMINI_MODEL to a model from your account, "
                "for example 'models/gemini-2.5-flash' or run the ModelService.ListModels to discover models."
            )
            raise HTTPException(status_code=502, detail={"error": "Gemini API 404", "body": error_body, "hint": hint})
        # Quota errors and other statuses should be surfaced clearly
        if status == 429:
            raise HTTPException(status_code=502, detail={"error": "Gemini API quota exceeded", "body": error_body})
        raise HTTPException(status_code=502, detail=f"Gemini API error: {error_body}")
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"Chat error: {str(exc)}")
