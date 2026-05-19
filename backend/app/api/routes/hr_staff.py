from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, Response, UploadFile
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from app.crud.hr_staff import (
    create_staff,
    create_attachment,
    delete_staff,
    export_staff_rows,
    get_attachment,
    get_staff,
    list_attachments,
    list_attachments_for_staff_ids,
    list_staff,
    update_staff,
)
from app.core.config import settings
from app.db.session import get_db
from app.schemas.hr_staff_attachment import HRStaffAttachmentRead
from app.schemas.hr_staff import HRStaffCreate, HRStaffRead, HRStaffUpdate

router = APIRouter(prefix="/staff")


@router.post("", response_model=HRStaffRead)
def create(payload: HRStaffCreate, db: Session = Depends(get_db)):
    return create_staff(db, payload)


@router.get("", response_model=list[HRStaffRead])
def list_(
    db: Session = Depends(get_db),
    skip: int = 0,
    limit: int = Query(default=50, le=500),
    q: Optional[str] = None,
    district: Optional[str] = None,
    designation: Optional[str] = None,
    facility_name: Optional[str] = None,
    employment_type: Optional[str] = None,
):
    return list_staff(
        db,
        skip=skip,
        limit=limit,
        q=q,
        district=district,
        designation=designation,
        facility_name=facility_name,
        employment_type=employment_type,
    )


@router.get("/export")
def export(
    request: Request,
    db: Session = Depends(get_db),
    format: str = Query(default="xlsx", pattern="^(csv|xlsx)$"),
    q: Optional[str] = None,
    district: Optional[str] = None,
    designation: Optional[str] = None,
    facility_name: Optional[str] = None,
    employment_type: Optional[str] = None,
):
    rows = export_staff_rows(
        db,
        q=q,
        district=district,
        designation=designation,
        facility_name=facility_name,
        employment_type=employment_type,
    )

    if format == "csv":
        fieldnames = [
            "id",
            "full_name",
            "gender",
            "date_of_birth",
            "designation",
            "cadre",
            "employment_type",
            "phone",
            "email",
            "facility_name",
            "facility_type",
            "district",
            "block",
            "posting_place",
            "date_of_joining",
            "remarks",
            "extra",
            "created_at",
            "updated_at",
        ]
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

        content = output.getvalue()
        headers = {"Content-Disposition": 'attachment; filename="hr_staff.csv"'}
        return Response(content=content, media_type="text/csv", headers=headers)

    # XLSX (formatted) export
    wb = Workbook()
    ws = wb.active
    ws.title = "Staff"

    base_url = str(request.base_url).rstrip("/")

    staff_headers = [
        "ID",
        "Full Name",
        "Gender",
        "DOB",
        "Designation",
        "Cadre",
        "Employment Type",
        "Phone",
        "Email",
        "Facility Name",
        "Facility Type",
        "District",
        "Block",
        "Posting Place",
        "Date of Joining",
        "Remarks",
        "Attachments",
        "Created",
        "Updated",
    ]
    ws.append(staff_headers)
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(staff_headers))}1"

    staff_ids = [int(r["id"]) for r in rows]
    atts = list_attachments_for_staff_ids(db, staff_ids)
    atts_by_staff: dict[int, list] = {}
    for a in atts:
        atts_by_staff.setdefault(a.staff_id, []).append(a)

    for r in rows:
        staff_id = int(r["id"])
        attachment_list_url = f"{base_url}/api/hr/staff/{staff_id}/attachments"
        attachment_count = len(atts_by_staff.get(staff_id, []))
        attachment_text = f"View ({attachment_count})" if attachment_count else ""

        ws.append(
            [
                staff_id,
                r.get("full_name", ""),
                r.get("gender", ""),
                r.get("date_of_birth", ""),
                r.get("designation", ""),
                r.get("cadre", ""),
                r.get("employment_type", ""),
                r.get("phone", ""),
                r.get("email", ""),
                r.get("facility_name", ""),
                r.get("facility_type", ""),
                r.get("district", ""),
                r.get("block", ""),
                r.get("posting_place", ""),
                r.get("date_of_joining", ""),
                r.get("remarks", ""),
                attachment_text,
                r.get("created_at", ""),
                r.get("updated_at", ""),
            ]
        )

        # Add hyperlink to attachments list
        if attachment_text:
            cell = ws.cell(row=ws.max_row, column=17)
            cell.hyperlink = attachment_list_url
            cell.style = "Hyperlink"

    # Column sizing (simple heuristic)
    for col in range(1, len(staff_headers) + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 18
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["J"].width = 26
    ws.column_dimensions["P"].width = 30
    ws.column_dimensions["Q"].width = 14

    # Attachments sheet (each row has direct download hyperlink)
    ws2 = wb.create_sheet(title="Attachments")
    att_headers = ["Staff ID", "Attachment ID", "File Name", "Content Type", "Created", "Download"]
    ws2.append(att_headers)
    for cell in ws2[1]:
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(att_headers))}1"

    for a in sorted(atts, key=lambda x: (x.staff_id, x.id)):
        download_url = f"{base_url}/api/hr/staff/{a.staff_id}/attachments/{a.id}"
        ws2.append([a.staff_id, a.id, a.original_filename, a.content_type or "", str(a.created_at), "Download"])
        dcell = ws2.cell(row=ws2.max_row, column=6)
        dcell.hyperlink = download_url
        dcell.style = "Hyperlink"

    for col in range(1, len(att_headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 24
    ws2.column_dimensions["C"].width = 40
    ws2.column_dimensions["F"].width = 16

    out = io.BytesIO()
    wb.save(out)
    content = out.getvalue()
    headers = {"Content-Disposition": 'attachment; filename="hr_staff.xlsx"'}
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.post("/{staff_id}/attachments", response_model=HRStaffAttachmentRead)
def upload_attachment(
    staff_id: int,
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    staff = get_staff(db, staff_id)
    if staff is None:
        raise HTTPException(status_code=404, detail="Staff record not found")

    att = create_attachment(
        db,
        staff_id=staff_id,
        original_filename=file.filename or "attachment",
        content_type=file.content_type,
    )

    uploads_root = Path(settings.uploads_dir)
    uploads_root.mkdir(parents=True, exist_ok=True)
    stored_path = uploads_root / att.stored_filename

    with stored_path.open("wb") as f:
        while True:
            chunk = file.file.read(1024 * 1024)
            if not chunk:
                break
            f.write(chunk)

    base_url = str(request.base_url).rstrip("/")
    att_read = HRStaffAttachmentRead.model_validate(att)
    att_read.url = f"{base_url}/api/hr/staff/{staff_id}/attachments/{att.id}"
    return att_read


@router.get("/{staff_id}/attachments", response_model=list[HRStaffAttachmentRead])
def list_staff_attachments(staff_id: int, request: Request, db: Session = Depends(get_db)):
    staff = get_staff(db, staff_id)
    if staff is None:
        raise HTTPException(status_code=404, detail="Staff record not found")

    base_url = str(request.base_url).rstrip("/")
    items = list_attachments(db, staff_id)
    out: list[HRStaffAttachmentRead] = []
    for a in items:
        r = HRStaffAttachmentRead.model_validate(a)
        r.url = f"{base_url}/api/hr/staff/{staff_id}/attachments/{a.id}"
        out.append(r)
    return out


@router.get("/{staff_id}/attachments/{attachment_id}")
def download_attachment(staff_id: int, attachment_id: int, db: Session = Depends(get_db)):
    att = get_attachment(db, staff_id, attachment_id)
    if att is None:
        raise HTTPException(status_code=404, detail="Attachment not found")

    stored_path = Path(settings.uploads_dir) / att.stored_filename
    if not stored_path.exists():
        raise HTTPException(status_code=404, detail="Attachment file missing on server")

    return FileResponse(
        path=str(stored_path),
        media_type=att.content_type or "application/octet-stream",
        filename=att.original_filename,
    )


@router.get("/{staff_id}", response_model=HRStaffRead)
def get_one(staff_id: int, db: Session = Depends(get_db)):
    staff = get_staff(db, staff_id)
    if staff is None:
        raise HTTPException(status_code=404, detail="Staff record not found")
    return staff


@router.patch("/{staff_id}", response_model=HRStaffRead)
def patch(staff_id: int, payload: HRStaffUpdate, db: Session = Depends(get_db)):
    staff = update_staff(db, staff_id, payload)
    if staff is None:
        raise HTTPException(status_code=404, detail="Staff record not found")
    return staff


@router.delete("/{staff_id}")
def delete(staff_id: int, db: Session = Depends(get_db)):
    ok = delete_staff(db, staff_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Staff record not found")
    return {"status": "deleted"}
